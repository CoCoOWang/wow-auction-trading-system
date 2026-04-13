import openpyxl
from .utils import get_db_connection, is_sheet_processed, mark_sheet_done, ensure_session, load_fee_params
from datetime import datetime
from .logger import logger
import re

sheet_parsing_status = True
file_parsing_status = True

# 可选：是否保留 auction_detail_data 全量深度（默认 False=只做本期缓存）
STORE_FULL_DETAIL_DEPTH = False


# 兼容：5堆叠20 / 5 堆叠 20 / 5x20 / 5×20 / 5*20
_STACK_RE_CN = re.compile(r'(\d+)\s*堆叠\s*(\d+)', re.I)
_STACK_RE_X  = re.compile(r'(\d+)\s*[xX×*]\s*(\d+)')


def parse_stack_size(qty_field) -> int:
    """
    从“可购买”字段里解析出 y（每堆数量）。
    - 匹配 'x堆叠y' 优先；否则尝试 'x×y' / 'x*y' / 'x x y'
    - 解析失败返回 1（视作不可堆叠）
    """
    if qty_field is None:
        return 1
    s = str(qty_field).strip()
    m = _STACK_RE_CN.search(s)
    if m:
        return int(m.group(2))
    m = _STACK_RE_X.search(s)
    if m:
        return int(m.group(2))
    # 最后再兜底：纯数字或其他形态都当 1
    return 1


def ensure_catalog_row(cursor, name, quality, item_string, timestamp, catalog_known: set):
    """
    确保 auction_item_list 里存在该物品：
    - 首次插入：is_stackable=0, max_stack_size=NULL（不要碰到将来可堆叠的升级）
    - 已存在：只更新 name/quality（不动 is_stackable/max_stack_size）
    """
    if item_string in catalog_known:
        # 已有目录行：只确保基础信息一致
        cursor.execute("""
            UPDATE auction_item_list
               SET name=%s, quality=%s
             WHERE itemString=%s
        """, (name, quality, item_string))
        return

    cursor.execute("""
        INSERT INTO auction_item_list
            (name, quality, itemString, createTime, is_stackable, max_stack_size)
        VALUES (%s,    %s,      %s,          %s,         0,           NULL)
        ON DUPLICATE KEY UPDATE
            name=VALUES(name),
            quality=VALUES(quality)
    """, (name, quality, item_string, timestamp))
    catalog_known.add(item_string)


def _upsert_seller_stats(cursor, item_string, seller, create_time, listings=0, takedowns=0, transactions=0):
    stat_date = create_time.split(" ")[0]
    cursor.execute("""
        INSERT INTO seller_stats_daily(stat_date, itemString, seller, listings, takedowns, transactions)
        VALUES (%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            listings = listings + VALUES(listings),
            takedowns = takedowns + VALUES(takedowns),
            transactions = transactions + VALUES(transactions)
    """, (stat_date, item_string, seller, int(listings), int(takedowns), int(transactions)))


class AuctionCurrent:
    def __init__(self, item_id, name, price, is_owned, total_quantity, time_left, seller, quality, item_string, create_time):
        self.itemId = item_id
        self.name = name
        self.price = price
        self.isOwned = is_owned
        self.totalQuantity = total_quantity
        self.timeLeft = time_left
        self.seller = seller
        self.quality = quality
        self.itemString = item_string
        self.createTime = create_time


class AuctionDetail:
    def __init__(self, item_id, name, unit_price, total_price, quantity, time_left, seller, quality, item_string):
        self.itemId = item_id
        self.name = name
        self.unitPrice = unit_price
        self.totalPrice = total_price
        self.quantity = quantity
        self.timeLeft = time_left
        self.seller = seller
        self.quality = quality
        self.itemString = item_string
        self.hasExisted = False  # 默认值为 False


class AuctionItemList:
    def __init__(self, name, quality, item_string, create_time):
        self.name = name
        self.quality = quality
        self.itemString = item_string
        self.createTime = create_time


# 读取数据库数据
def load_mysql_data():
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        query_current = "SELECT itemId, name, price, isOwned, totalQuantity, timeLeft, seller, quality, itemString, createTime FROM current_auction_data"
        query_detail = "SELECT itemId, name, unitPrice, totalPrice, quantity, timeLeft, seller, quality, itemString FROM auction_detail_data"
        query_item_list = "SELECT name, quality, itemString, createTime FROM auction_item_list"

        cursor.execute(query_current)
        data_current = cursor.fetchall()

        cursor.execute(query_detail)
        data_detail = cursor.fetchall()

        cursor.execute(query_item_list)
        data_item_list = cursor.fetchall()

        # 关闭数据库连接
        cursor.close()
        conn.close()

        # 转换为普通对象数组
        df_current = [AuctionCurrent(*row) for row in data_current]
        df_detail = [AuctionDetail(*row) for row in data_detail]
        df_item_list = [AuctionItemList(*row) for row in data_item_list]

        return df_current, df_detail, df_item_list

    except Exception as e:
        print(f"❌ load_mysql_data 数据查询失败: {e}")
        return None, None


def calculate_unit_price(total_price_str, quantity_str, status):
    # 解析输入
    total_price = int(total_price_str)  # 总价（铜）
    if status:
        quantity = int(quantity_str)
    else:
        parts = quantity_str.split("堆叠")
        part2 = int(parts[1])
        quantity = part2  # 物品数量

    # 计算单价（铜）
    unit_price = total_price // quantity

    # 返回字符串格式的铜单位单价
    return str(unit_price)


def get_total_quantity(quantity_str):
    try:
        parts = quantity_str.split("堆叠")
        part1 = int(parts[0].strip())
        part2 = int(parts[1].strip())
        return part1 * part2
    except (IndexError, ValueError) as e:
        print(f"⚠️ 格式异常: {quantity_str} -> {e}")
        return 0  # 或其他默认值


def get_next_and_further_time_left(current_time_left: str):
    time_mapping = {
        "8h-24h": ("2h-8h", "30m-2h"),
        "2h-8h": ("30m-2h", "<30m"),
        "30m-2h": ("<30m", ""),
        "<30m": ("", "")
    }

    # 获取 `next` 和 `further`，如果 `current_time_left` 不存在，则返回 (False, False)
    return time_mapping.get(current_time_left, ("", ""))


def time_left_probability(previous_time_left: str, delta_hours: float):
    """
    计算过去 delta_hours 小时后，物品在各个剩余时间区间的概率
    """
    time_intervals = {
        "8h-24h": (8, 24),
        "2h-8h": (2, 8),
        "30m-2h": (0.5, 2),
        "<30m": (0, 0.5),
    }

    if previous_time_left not in time_intervals:
        return "未知时间区间"

    min_time, max_time = time_intervals[previous_time_left]

    # 计算进入下一个区间的概率
    if delta_hours < max_time / 4:  # 时间过去较少，大概率仍在当前区间
        return {previous_time_left: 0.85, "Next": 0.15, "Further": 0}
    elif delta_hours < max_time / 3:  # 可能进入下一个区间
        return {previous_time_left: 0.6, "Next": 0.35, "Further": 0.05}
    elif delta_hours < max_time * 0.75:  # 过了一大半的时间，可能性变化
        return {previous_time_left: 0.3, "Next": 0.5, "Further": 0.2}
    else:  # 超过最大时间，100% 进入下一个区间
        return {previous_time_left: 0, "Next": 0.8, "Further": 0.2}


def get_hour_difference(time1: str, time2: str):
    """
    计算两个时间（yyyy-MM-dd HH:mm:ss 格式）之间的小时差
    - 有小数部分时，保留 1 位小数
    - 没有小数部分时，返回整数
    """
    fmt = "%Y-%m-%d %H:%M:%S"
    dt1 = datetime.strptime(time1, fmt)
    dt2 = datetime.strptime(time2, fmt)

    # 计算时间差（小时）
    delta_hours = abs((dt2 - dt1).total_seconds()) / 3600

    # 如果是整数，返回 int，否则保留 1 位小数
    return int(delta_hours) if delta_hours.is_integer() else round(delta_hours, 1)


# 计算 diff
def compare_data(sheet, header_map, df_current, df_detail, df_item_list, timestamp, time_interval):
    results = []

    # 当前的最近单价
    current_unit_price = ""
    old_detail = None
    # 当前的物品名称
    current_name = ""
    previous_unit_price = None
    # 记录本轮新详情数据
    new_detail_data = []
    # 旧数据当前数据
    old_data = None
    # 当前的物品品质
    current_quality = None
    # 当前的物品标识符
    current_item_string = None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 判断当前有没有物品变更
        has_changed = False
        item_id = row[header_map["物品ID"]]
        name = row[header_map["名称"]]
        # 新数据-总价格
        new_total_price = row[header_map["价格"]]
        # 新数据-总库存
        new_total_quantity = row[header_map["可购买"]]
        # 新数据-卖家
        new_seller = row[header_map["卖家"]]
        # 新数据-品质
        quality = row[header_map["品质"]]
        # 新数据-标识符
        item_string = row[header_map["标识符"]]
        # 旧数据
        if name:
            old_data = next((item for item in df_current if item.itemId == item_id and item.name == name and item.quality == quality and item.itemString == item_string), None)

        if not old_data:
            if name is None:
                # 新增一种商品
                results.append((item_id, current_name, new_seller, 4, calculate_unit_price(new_total_price, new_total_quantity, False), get_total_quantity(new_total_quantity), 0,
                                0, 0, current_quality, current_item_string, timestamp))
            else:
                current_name = name
                current_quality = quality
                current_item_string = item_string
                is_newly_item = next((item for item in df_item_list if item.itemString == item_string), None)
                if is_newly_item is None:
                    results.append((item_id, name, new_seller, 5, 0, 0, 0, 0, 0, quality, item_string, timestamp))
            continue

        if name is None:
            new_data_processed_status = True
            processed_total_quantity = get_total_quantity(new_total_quantity)

            # 新数据-单价
            new_unit_price = calculate_unit_price(new_total_price, new_total_quantity, False)
            # 新数据-剩余时间
            new_time_left = row[header_map["剩余时间"]]

            # 新数据是否是新增数据
            is_new_added = True

            for detail_row in old_detail:
                if not detail_row.hasExisted:
                    # 旧数据-详情-总库存
                    old_detail_quantity = detail_row.quantity
                    # 旧数据-详情-单价
                    old_detail_unit_price = detail_row.unitPrice
                    # 旧数据-详情-剩余时间
                    old_detail_time_left = detail_row.timeLeft
                    # 旧数据-详情-卖家
                    old_detail_seller = detail_row.seller

                    processed_previous_quantity = get_total_quantity(old_detail_quantity)

                    # 判断新旧是否是同一条数据
                    if new_seller == old_detail_seller and new_unit_price == old_detail_unit_price and new_time_left == old_detail_time_left:
                        is_new_added = False

                        detail_row.hasExisted = True

                        if processed_total_quantity != processed_previous_quantity:
                            has_changed = True
                            if processed_total_quantity > processed_previous_quantity:
                                # 代表着上架了商品（其实判断不了上架商品中被出售的那一部分）
                                results.append((item_id, current_name, new_seller, 0, new_unit_price, processed_total_quantity, 0, processed_total_quantity - processed_previous_quantity, 0, current_quality, current_item_string, timestamp))
                            else:
                                # 代表着下架或者交易
                                # 当前的商品的最低单价高于前一个的最低单价
                                if current_unit_price > previous_unit_price:
                                    if int(new_unit_price) <= current_unit_price:
                                        # 大概率是成交了
                                        results.append((item_id, current_name, new_seller, 0, new_unit_price,
                                                        processed_total_quantity, 0, 0,
                                                        processed_previous_quantity - processed_total_quantity,
                                                        current_quality,
                                                        current_item_string,
                                                        timestamp))
                                    else:
                                        # 大概率是下架了
                                        results.append((item_id, current_name, new_seller, 0, new_unit_price,
                                                        processed_total_quantity,
                                                        processed_previous_quantity - processed_total_quantity, 0, 0,
                                                        current_quality,
                                                        current_item_string,
                                                        timestamp))
                                else:
                                    if int(new_unit_price) <= previous_unit_price:
                                        # 大概率是成交了
                                        results.append((item_id, current_name, new_seller, 0, new_unit_price,
                                                        processed_total_quantity, 0, 0,
                                                        processed_previous_quantity - processed_total_quantity,
                                                        current_quality,
                                                        current_item_string,
                                                        timestamp))
                                    else:
                                        # 大概率是下架了
                                        results.append((item_id, current_name, new_seller, 0, new_unit_price,
                                                        processed_total_quantity,
                                                        processed_previous_quantity - processed_total_quantity, 0, 0,
                                                        current_quality,
                                                        current_item_string,
                                                        timestamp))
                        break
            #  接下来判断是否是剩余时间进入了下一个档位（需要自动化判断是否可以在维护后，最先登录）
            if is_new_added:
                for detail_row in old_detail:
                    if not detail_row.hasExisted:
                        # 旧数据-详情-总库存
                        old_detail_quantity = detail_row.quantity
                        # 旧数据-详情-单价
                        old_detail_unit_price = detail_row.unitPrice
                        # 旧数据-详情-卖家
                        old_detail_seller = detail_row.seller
                        # 旧数据-详情-剩余时间的下一个档位
                        probability_detail = time_left_probability(detail_row.timeLeft, time_interval)
                        next_time, further_time = get_next_and_further_time_left(detail_row.timeLeft)
                        probability_time_left = []
                        if probability_detail["Next"] > 0:
                            probability_time_left.append(next_time)
                        if probability_detail["Further"] > 0:
                            probability_time_left.append(further_time)

                        has_probability_status = False

                        processed_old_quantity = get_total_quantity(old_detail_quantity)

                        for probability_item in probability_time_left:
                            # 判断新旧是否是同一条数据
                            if new_seller == old_detail_seller and new_unit_price == old_detail_unit_price and new_time_left == probability_item:
                                is_new_added = False
                                detail_row.hasExisted = True
                                if processed_total_quantity != processed_old_quantity:
                                    has_changed = True
                                    if processed_total_quantity > processed_old_quantity:
                                        # 代表着上架了商品（其实判断不了上架商品中被出售的那一部分）
                                        results.append((item_id, current_name, new_seller, 0, new_unit_price, processed_total_quantity, 0, processed_total_quantity - processed_old_quantity, 0, current_quality, current_item_string, timestamp))
                                    else:
                                        # 代表着下架或者交易
                                        # 当前的商品的最低单价高于前一个的最低单价
                                        if current_unit_price > previous_unit_price:
                                            if int(new_unit_price) <= current_unit_price:
                                                # 大概率是成交了
                                                results.append((item_id, current_name, new_seller, 0, new_unit_price, processed_total_quantity, 0, 0,
                                                            processed_old_quantity - processed_total_quantity, current_quality, current_item_string, timestamp))
                                            else:
                                                # 大概率是下架了
                                                results.append((item_id, current_name, new_seller, 0, new_unit_price,
                                                                processed_total_quantity,
                                                                processed_old_quantity - processed_total_quantity, 0, 0,
                                                                current_quality,
                                                                current_item_string,
                                                                timestamp))
                                        else:
                                            if int(new_unit_price) <= previous_unit_price:
                                                # 大概率是成交了
                                                results.append((item_id, current_name, new_seller, 0, new_unit_price,
                                                                processed_total_quantity, 0, 0,
                                                                processed_old_quantity - processed_total_quantity,
                                                                current_quality,
                                                                current_item_string,
                                                                timestamp))
                                            else:
                                                # 大概率是下架了
                                                results.append((item_id, current_name, new_seller, 0, new_unit_price,
                                                                processed_total_quantity,
                                                                processed_old_quantity - processed_total_quantity, 0, 0,
                                                                current_quality,
                                                                current_item_string,
                                                                timestamp))

                                has_probability_status = True
                                break

                        if has_probability_status:
                            break
            # 判断是新增数据
            if is_new_added:
                results.append((item_id, current_name, new_seller, 1, new_unit_price, processed_total_quantity, 0, 0, 0, current_quality, current_item_string, timestamp))
                has_changed = True
                new_data_processed_status = False

            # 判断是没有变更的数据
            if not has_changed:
                results.append((item_id, current_name, new_seller, 3, new_unit_price, processed_total_quantity, 0, 0, 0, current_quality, current_item_string, timestamp))

            new_detail_data.append({'price': new_unit_price, 'quantity': processed_total_quantity, 'seller': new_seller,
                                    'time_left': new_time_left, 'status': new_data_processed_status})
        else:
            # 判断old_detail有值的情况下，分析哪些数据是下架或者交易数据（消失状态）
            if old_detail is not None:
                for detail_row in old_detail:
                    if not detail_row.hasExisted:
                        processed_detail_quantity = get_total_quantity(detail_row.quantity)
                        is_it_taken_down = False
                        # 当前的商品的最低单价高于前一个的最低单价
                        if current_unit_price > previous_unit_price:
                            if int(detail_row.unitPrice) <= current_unit_price:
                                if detail_row.timeLeft == "<30m":
                                    # 大概率下架了
                                    is_it_taken_down = True
                            else:
                                # 大概率下架了
                                is_it_taken_down = True
                        else:
                            if int(detail_row.unitPrice) <= previous_unit_price:
                                is_it_automatically_taken_down = True
                                process_first_step = False
                                process_second_step = False
                                for new_detail_item in new_detail_data:
                                    price = new_detail_item['price']
                                    quantity = new_detail_item['quantity']
                                    seller = new_detail_item['seller']
                                    status = new_detail_item['status']
                                    if not status:
                                        # 如果新的商品数据列表内存在与旧数据的卖家、商品单价相同、商品数量相同的数据（卖家调整了更长的出售时间）
                                        if seller == detail_row.seller and price == detail_row.unitPrice and quantity == processed_detail_quantity:
                                            new_detail_item['status'] = True
                                            # 大概率下架了
                                            is_it_taken_down = True
                                            process_first_step = True
                                            process_second_step = True
                                            is_it_automatically_taken_down = False
                                            break
                                if not process_first_step:
                                    for new_detail_item in new_detail_data:
                                        price = new_detail_item['price']
                                        quantity = new_detail_item['quantity']
                                        seller = new_detail_item['seller']
                                        time_left = new_detail_item['time_left']
                                        status = new_detail_item['status']
                                        if not status:
                                            # 如果新的商品数据列表内存在与旧数据的卖家、商品数量相同、剩余时间相同的数据（卖家调整了商品单价）
                                            if seller == detail_row.seller and quantity == processed_detail_quantity and time_left == detail_row.timeLeft:
                                                # 如果新的商品的单价小于detail的商品单价
                                                if price < detail_row.unitPrice:
                                                    new_detail_item['status'] = True
                                                    # 大概率下架了
                                                    is_it_taken_down = True
                                                    process_second_step = True
                                                    is_it_automatically_taken_down = False
                                                    break
                                if not process_second_step:
                                    for new_detail_item in new_detail_data:
                                        price = new_detail_item['price']
                                        quantity = new_detail_item['quantity']
                                        seller = new_detail_item['seller']
                                        status = new_detail_item['status']
                                        if not status:
                                            # 如果新的商品数据列表内存在与旧数据的卖家相同的数据（卖家可能调整了商品单价、也可能调整了剩余时间）
                                            if seller == detail_row.seller:
                                                # 如果新的商品的单价小于等于detail的商品单价
                                                if price <= detail_row.unitPrice:
                                                    new_detail_item['status'] = True
                                                    # 大概率下架了
                                                    is_it_taken_down = True
                                                    is_it_automatically_taken_down = False
                                                    # 如果新的商品数量小于detail的商品数量
                                                    if quantity < processed_detail_quantity:
                                                        # 成交了（detail的商品数量 - 新的商品数量）个商品，成交价格为新的商品的价格
                                                        # 大概率是成交了
                                                        results.append(
                                                            (detail_row.itemId, current_name, detail_row.seller, 2,
                                                             price,
                                                             (processed_detail_quantity - quantity), 0, 0,
                                                             (processed_detail_quantity - quantity),
                                                             detail_row.quality,
                                                             detail_row.itemString,
                                                             timestamp))
                                                        # 上架了（detail的商品数量 - 新的商品数量）个商品，上架价格为新的商品的价格
                                                        results.append(
                                                            (detail_row.itemId, current_name, detail_row.seller, 1,
                                                             price,
                                                             (processed_detail_quantity - quantity), 0,
                                                             (processed_detail_quantity - quantity),
                                                             0,
                                                             detail_row.quality,
                                                             detail_row.itemString,
                                                             timestamp))
                                                    break

                                # 1）和2）和3）都不满足时
                                if is_it_automatically_taken_down:
                                    # 如果当前商品剩余时间是<30m
                                    if detail_row.timeLeft == "<30m":
                                        # 大概率下架了
                                        is_it_taken_down = True
                            else:
                                # 大概率下架了
                                is_it_taken_down = True

                        if is_it_taken_down:
                            # 大概率下架了
                            results.append((detail_row.itemId, current_name, detail_row.seller, 2, detail_row.unitPrice,
                                            processed_detail_quantity, processed_detail_quantity, 0, 0,
                                            detail_row.quality, detail_row.itemString, timestamp))
                        else:
                            # 大概率是成交了
                            results.append(
                                (detail_row.itemId, current_name, detail_row.seller, 2, detail_row.unitPrice,
                                 processed_detail_quantity, 0, 0, processed_detail_quantity, detail_row.quality,
                                 detail_row.itemString, timestamp))
            old_detail = [item for item in df_detail if item.itemId == item_id and item.name == name and item.quality == quality and item.itemString == item_string]

            current_name = row[header_map["名称"]]
            current_quality = row[header_map["品质"]]
            current_item_string = row[header_map["标识符"]]
            # 当前的最低单价
            current_unit_price = int(new_total_price)
            # 上一次的最低单价
            previous_unit_price = int(old_data.price)
            # 记录本次新详情数据
            new_detail_data = []

    return results


def parse_sheet_data(file_path, timestamp, has_stopped, sheet_name):
    """
    解析单个 Sheet 并入库（可累加 + 幂等 + 会话 + 可选保留深度）
    """
    global sheet_parsing_status, file_parsing_status
    conn = None  # 先初始化

    try:
        sheet_parsing_status = False
        if not has_stopped:
            conn = get_db_connection()
            cursor = conn.cursor()

            # —— 加载“已知目录物品”（存在即不再 INSERT）
            cursor.execute("SELECT itemString FROM auction_item_list")
            catalog_known = set(r[0] for r in cursor.fetchall())

            # —— 新增：加载已标记可堆叠的物品，避免重复判断/更新
            cursor.execute("SELECT itemString FROM auction_item_list WHERE is_stackable=1")
            stackable_known = set(row[0] for row in cursor.fetchall())

            # 若想“只判一次”，保持 True；若想后续遇到更大的 y 时更新 max_stack_size，改为 False
            ONLY_SET_ONCE = True

            # 1) 会话登记 & 进度判断
            ensure_session(conn, file_path)
            if is_sheet_processed(conn, file_path, sheet_name):
                logger.info(f"已处理过：{file_path} / {sheet_name}（跳过）")
                sheet_parsing_status = True
                return

            # 2) 载入 Excel
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]

            # 读取标题行
            headers = next(sheet.iter_rows(values_only=True))

            # 将列名映射到索引（例如 {"姓名": 0, "年龄": 1, "城市": 2}）
            header_map = {header: idx for idx, header in enumerate(headers)}

            # 3) 判断是否有“上一期缓存”
            # 判断current_auction_data表有没有数据
            cursor.execute("""SELECT 1 FROM current_auction_data LIMIT 1""")
            cursor_status = cursor.fetchone() is not None

            # 4) 如有上一期，则做对比并幂等写入各事实表 + 卖家画像
            # current_auction_data有数据
            if cursor_status:
                # 获取当前表中全部数据
                df_current, df_detail, df_item_list = load_mysql_data()
                df_current_create_time = df_current[0].createTime
                time_interval = get_hour_difference(df_current_create_time.strftime("%Y-%m-%d %H:%M:%S"), timestamp)
                diffs = compare_data(sheet, header_map, df_current, df_detail, df_item_list, timestamp, time_interval)
                print(f"数据对比已完成！！！对比数据量是{len(diffs)}条")

                # 分析diffs向数据库中插入数据
                for item in diffs:
                    item_id, name, seller, status, unit_price, total_quantity, decline_listing_count, listing_count, transaction_count, quality, item_string, create_time = item
                    # 写入 listing/delisting/transaction/unchanged（幂等）
                    if status == 0 or status == 2:
                        if int(decline_listing_count) > 0:
                            cursor.execute("""
                                INSERT INTO delisting_detail_data (itemId, name, seller, unitPrice, totalQuantity, takenDownCount, status, quality, itemString, createTime)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                  totalQuantity=VALUES(totalQuantity),
                                  takenDownCount=VALUES(takenDownCount)
                            """, (item_id, name, seller, unit_price, total_quantity, decline_listing_count, status, quality, item_string, create_time))
                            _upsert_seller_stats(cursor, item_string, seller, create_time,
                                                 takedowns=decline_listing_count)
                        elif int(listing_count) > 0:
                            cursor.execute("""
                                INSERT INTO listing_detail_data (itemId, name, seller, status, unitPrice, totalQuantity, listingsCount, quality, itemString, createTime)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                  totalQuantity=VALUES(totalQuantity),
                                  listingsCount=VALUES(listingsCount)
                            """,(item_id, name, seller, 0, unit_price, total_quantity, listing_count, quality, item_string, create_time))
                            _upsert_seller_stats(cursor, item_string, seller, create_time, listings=listing_count)
                        elif int(transaction_count) > 0:
                            cursor.execute("""
                                INSERT INTO transaction_detail_data (itemId, name, seller, unitPrice, totalQuantity, transactionQuantity, status, quality, itemString, createTime)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                  totalQuantity=VALUES(totalQuantity),
                                  transactionQuantity=VALUES(transactionQuantity)
                            """,(item_id, name, seller, unit_price, total_quantity, transaction_count, status, quality, item_string, create_time))
                            _upsert_seller_stats(cursor, item_string, seller, create_time,
                                                 transactions=transaction_count)
                    # 新加入/起始点
                    elif status == 1 or status == 4:
                        cursor.execute("""
                            INSERT INTO listing_detail_data (itemId, name, seller, status, unitPrice, totalQuantity, listingsCount, quality, itemString, createTime)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                              totalQuantity=VALUES(totalQuantity),
                              listingsCount=VALUES(listingsCount)
                        """,(item_id, name, seller, status, unit_price, total_quantity, total_quantity, quality, item_string, create_time))
                        _upsert_seller_stats(cursor, item_string, seller, create_time, listings=total_quantity)
                    # 未变更
                    elif status == 3:
                        cursor.execute("""
                            INSERT INTO unchaged_detail_data (itemId, name, seller, unitPrice, totalQuantity, quality, itemString, createTime)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE totalQuantity=VALUES(totalQuantity)
                        """, (item_id, name, seller, unit_price, total_quantity, quality, item_string, create_time))
                    # 新物品入目录
                    elif status == 5:
                        cursor.execute("""
                            INSERT INTO auction_item_list (name, quality, itemString, createTime)
                            VALUES (%s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE quality=VALUES(quality)
                        """, (name, quality, item_string, create_time))

                # 删除current_auction_data表、auction_detail_data数据
                # 要操作的表名
                tables = ["current_auction_data", "auction_detail_data"]
                for table in tables:
                    # 清空表数据
                    cursor.execute(f"DELETE FROM {table};")
                    # 重置自增主键 ID
                    cursor.execute(f"ALTER TABLE {table} AUTO_INCREMENT = 1;")
                    print(f"表 {table} 已清空，并将自增 ID 置为 1")

            # 5) 写入“本期快照”：auction_history_data + current_auction_data + auction_detail_data（可选保留）
            inserted_rows = 0
            current_name = ""
            current_quality = 0
            current_item_string = ""

            # 插入新的sheet页的数据
            for row in sheet.iter_rows(min_row=2, values_only=True):

                item_id = row[header_map["物品ID"]]
                name = row[header_map["名称"]]
                price = row[header_map["价格"]]
                is_owned = row[header_map["我的售品？"]]
                total_quantity = row[header_map["可购买"]]
                time_left = row[header_map["剩余时间"]]
                seller = row[header_map["卖家"]]
                quality = row[header_map["品质"]]
                item_string = row[header_map["标识符"]]

                if name is None:
                    cursor.execute("""
                        INSERT INTO auction_detail_data (itemId, name, unitPrice, totalPrice, quantity, timeLeft, seller, quality, itemString)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (item_id, current_name, calculate_unit_price(price, total_quantity, False), price, total_quantity, time_left, seller, current_quality, current_item_string))
                    # 新的（精确从“x堆叠y”里取 y；只判一次）
                    stack_y = parse_stack_size(total_quantity)  # y = 每堆数量
                    if stack_y > 1:
                        if current_item_string not in stackable_known:
                            # 第一次发现该物品可堆叠 → 只记录一次
                            cursor.execute("""
                                INSERT INTO auction_item_list (name, quality, itemString, createTime, is_stackable, max_stack_size)
                                VALUES (%s,%s,%s,%s,1,%s)
                                ON DUPLICATE KEY UPDATE
                                  is_stackable = 1,
                                  max_stack_size = IFNULL(max_stack_size, VALUES(max_stack_size))
                            """, (current_name, current_quality, current_item_string, timestamp, stack_y))
                            stackable_known.add(current_item_string)
                        else:
                            if not ONLY_SET_ONCE:
                                # 若允许后续更新更大的 y，这里再提升 max_stack_size
                                cursor.execute("""
                                    UPDATE auction_item_list
                                    SET max_stack_size = GREATEST(IFNULL(max_stack_size,0), %s)
                                    WHERE itemString = %s
                                """, (stack_y, current_item_string))

                else:
                    current_name = name
                    current_quality = quality
                    current_item_string = item_string
                    cursor.execute("""
                        INSERT INTO current_auction_data (itemId, name, price, isOwned, totalQuantity, timeLeft, seller, quality, itemString, createTime)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (item_id, name, price, is_owned, total_quantity, time_left, seller, quality, item_string, timestamp))
                    cursor.execute("""
                        INSERT INTO auction_history_data (itemId, name, unitPrice, totalQuantity, quality, itemString, createTime)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                          unitPrice=VALUES(unitPrice),
                          totalQuantity=VALUES(totalQuantity),
                          quality=VALUES(quality)
                    """, (item_id, name, price, total_quantity, quality, item_string, timestamp))
                    # 物品头行：先确保目录有条目（不可堆叠默认值）
                    ensure_catalog_row(cursor, name, quality, item_string, timestamp, catalog_known)

                inserted_rows += 1

            # 6) 进度落账
            mark_sheet_done(conn, file_path, sheet_name, timestamp, inserted_rows)
            conn.commit()
            print(f"Sheet {sheet_name} 数据存入数据库，时间戳: {timestamp}")
        else:
            # 文件解析完毕
            file_parsing_status = True

    except Exception as e:
        print(f"解析 Sheet 失败: {e}")
        if conn:
            conn.rollback()
    finally:
        sheet_parsing_status = True
        # 关闭连接
        if conn:
            cursor.close()
            conn.close()